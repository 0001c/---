import os
import shutil
from openpyxl import load_workbook
from tqdm import tqdm

import read
import ai_tool

def get_data():
    return read.main()

def ai_analysis(data):
    data_analysis = data.copy()
    data_str = str(data)
    data_response = ai_tool.main(data_str)
    data_analysis += data_response
    return data_analysis

def main_excel(data_analysis):
    try:
        template_file = "模板.xlsx"
        if not os.path.exists(template_file):
            raise FileNotFoundError(f"模板文件 {template_file} 不存在")
        
        output_file = f".\\output\\{data_analysis[0]}_xx调剂定制班1v1个人评估表.xlsx"
        shutil.copy2(template_file, output_file)
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        cells_to_modify = ['A1', 'A3', 'A5', 'A7', 'A9']
        data_to_write = [
            f"{data_analysis[0]}调剂整体规划",
            data_analysis[1],
            data_analysis[2],
            data_analysis[3],
            data_analysis[4]
        ]
        
        for cell, value in zip(cells_to_modify, data_to_write):
            ws[cell] = value
        
        wb.save(output_file)
        # 把这里的 print 去掉或者改用 pbar.write，避免打断 tqdm 进度条的显示
        # print(f"已成功保存: {output_file}") 
        
    except Exception as e:
        print(f"生成 Excel 时发生错误: {e}")

if __name__ == "__main__":
    # 1. 获取所有数据
    dist = get_data()
    if not dist:
        print("未获取到数据，程序结束。")
        exit()

    # ================= 新增：手动选择逻辑 =================
    print("\n" + "="*30)
    print("读取到的考生列表：")
    
    # 建立一个便于用户选择的映射字典
    # 使用 1, 2, 3... 作为显示序号，对应真实的源数据 key 和 value
    candidates = {}
    for i, (original_key, row_data) in enumerate(dist.items(), 1):
        # 假设 data1 是姓名 (根据你之前的代码逻辑，row[6] 开始的第一列)
        name = row_data.get('data1', '未知姓名')
        candidates[i] = (original_key, row_data)
        print(f"[{i}] {name}")
    print("="*30)
    
    selected_data = {}
    
    while True:
        print("\n请选择要生成规划表的考生序号：")
        print("（输入数字并用逗号分隔，例如 '1,3,5'。或者输入 'all' 处理所有人）")
        choice = input("你的选择: ").strip().lower()
        
        if choice == 'all':
            selected_data = dist
            break
        elif choice:
            try:
                # 将用户输入的字符串 "1, 3" 拆分并转换为整数列表 [1, 3]
                selected_indices = [int(x.strip()) for x in choice.split(',')]
                valid = True
                
                # 校验用户输入的序号是否都存在
                for idx in selected_indices:
                    if idx not in candidates:
                        print(f"❌ 错误：序号 [{idx}] 不存在，请检查后重新输入！")
                        selected_data.clear() # 清空刚才可能部分加载的数据
                        valid = False
                        break
                    else:
                        # 如果有效，将真实的数据提取出来装入 selected_data
                        original_key, row_data = candidates[idx]
                        selected_data[original_key] = row_data
                
                if valid and selected_data:
                    break # 输入完全合法，跳出循环
            except ValueError:
                print("❌ 格式错误：请确保输入的是纯数字和半角逗号。")
        else:
            print("❌ 输入不能为空。")

    if not selected_data:
        print("未选择任何数据，程序结束。")
        exit()

    print(f"\n✅ 成功选择 {len(selected_data)} 位考生，开始准备处理...\n")
    # =====================================================

    # 2. 进度条与核心处理流程 (注意这里改为遍历 selected_data)
    total_steps = len(selected_data) * 3  
    
    with tqdm(total=total_steps, desc="整体处理进度", unit="步") as pbar:
        for data in selected_data.values():
            data_list = list(data.values())
            name = data_list[0]
            
            # 步骤2：准备数据...
            pbar.set_description(f"准备数据中: {name}")
            data_analysis = [
                name,
                f''' 
                1.目前学历：{data_list[1]}
                2.本科院校及专业：{data_list[2]}
                3.一志愿报考院校：{data_list[3]}
                4.一志愿报考专业代码和全称：{data_list[4]}
                5.一志愿报考学硕还是专硕：{data_list[5]}
                6.一志愿报考专业学习方式：{data_list[6]}
                7.初试总分：{data_list[7]}
                8.外语科目和分数（标明英语一/二/其他小语种）：{data_list[8]}
                9.政治分数：{data_list[9]}
                10.专业课一代码+全称+分数：{data_list[10]}
                11.专业课二代码+全称+分数：{data_list[11]}  
                12.专项计划：{data_list[12]}
                13.有无艺术大类下其他方向特长：无
                14.本人手机号：{data_list[13]}
                15.紧急联系人手机号（联系不到你时确保能收到及时通知）：{data_list[14]}  
                ''',
                f'''
                1.学习方式的调剂意向：{data_list[15]}
                2.学硕专硕的调剂意向：{data_list[16]}
                3.学校区域的调剂意向：{data_list[17]}
                4.学校等级的调剂意向：{data_list[18]}
                5.艺术大类的调剂意向：{data_list[19]}
                '''
            ]
            pbar.update(1)
            
            # 步骤3：AI分析
            pbar.set_description(f"AI分析中: {name}")
            data_analysis = ai_analysis(data_analysis)
            pbar.update(1)
            
            # 步骤4：生成表格
            pbar.set_description(f"生成表格: {name}")
            main_excel(data_analysis)
            pbar.update(1)
        
        pbar.set_description("全部处理完成！")