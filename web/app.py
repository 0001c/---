"""
xx调剂定制班 - Web 服务端
Flask + SocketIO 实现文件上传、数据解析、AI 分析、实时进度推送、Excel 生成与下载
"""

import os
import sys
import uuid
import shutil
import zipfile
import threading
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO, emit
from openpyxl import load_workbook

# 将项目根目录加入 sys.path，以便导入 read / ai_tool
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, PROJECT_ROOT)

import read as reader
import ai_tool

# ─── 路径配置 ────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── Flask 初始化 ────────────────────────────────────────────────────
app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'templates'),
    static_folder=os.path.join(BASE_DIR, 'static'),
)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB 上传限制
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# ─── 会话存储（简单内存字典，单用户场景足够） ───────────────────────────
sessions = {}  # session_id -> { stats_file, template_file, students, ... }


# ═══════════════════════════════════════════════════════════════════════
#  工具函数
# ═══════════════════════════════════════════════════════════════════════

def allowed_file(filename):
    return filename and filename.lower().endswith(('.xlsx', '.xls'))


def parse_stats_file(filepath):
    """从上传的统计表文件中解析考生数据，返回 {index: {data1..data20}} 字典"""
    import pandas as pd
    import re

    result = {}
    try:
        df = pd.read_excel(filepath)
        for index, row in df.iterrows():
            row_data = {}
            for i, value in enumerate(row[6:], 1):
                if i == 20 and isinstance(value, str) and "〖" in value:
                    pattern = r"〖(.*?)〗"
                    matches = re.findall(pattern, value)
                    if matches:
                        value = matches[0]
                row_data[f'data{i}'] = value
            result[index] = row_data
    except Exception as e:
        print(f"[ERROR] 解析统计表失败: {e}")
        return {}
    return result


def build_data_analysis(data):
    """将考生原始数据格式化为 AI 分析所需的列表（复用 edit.py 逻辑）"""
    data_list = list(data.values())
    name = data_list[0]
    data_analysis = [
        name,
        f'''
        1.目前学历：{data_list[1]}
        2.本科院校及专业：{data_list[2]}
        3.一志愿报考院校：{data_list[3]}
        4.一志愿报考专业代码和全称：{data_list[4]}
        5.一志愿报考学硕还是专硕：{data_list[5]}
        6.一志愿报考专业学习方式：{data_list[6]}
        7.初试总分：{data_list[7]}
        8.外语科目和分数（标明英语一/二/其他小语种）：{data_list[8]}
        9.政治分数：{data_list[9]}
        10.专业课一代码+全称+分数：{data_list[10]}
        11.专业课二代码+全称+分数：{data_list[11]}
        12.专项计划：{data_list[12]}
        13.有无艺术大类下其他方向特长：无
        14.本人手机号：{data_list[13]}
        15.紧急联系人手机号（联系不到你时确保能收到及时通知）：{data_list[14]}
        ''',
        f'''
        1.学习方式的调剂意向：{data_list[15]}
        2.学硕专硕的调剂意向：{data_list[16]}
        3.学校区域的调剂意向：{data_list[17]}
        4.学校等级的调剂意向：{data_list[18]}
        5.艺术大类的调剂意向：{data_list[19]}
        '''
    ]
    return data_analysis, name


def generate_excel(data_analysis, template_path, output_path):
    """复用 edit.py 的 Excel 生成逻辑"""
    try:
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        ws = wb.active

        cells = ['A1', 'A3', 'A5', 'A7', 'A9']
        values = [
            f"{data_analysis[0]}调剂整体规划",
            data_analysis[1],
            data_analysis[2],
            data_analysis[3],
            data_analysis[4],
        ]
        for cell, value in zip(cells, values):
            ws[cell] = value

        wb.save(output_path)
        return True, None
    except Exception as e:
        return False, str(e)


# ═══════════════════════════════════════════════════════════════════════
#  路由
# ═══════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_files():
    """上传统计表和/或模板文件"""
    session_id = str(uuid.uuid4())[:8]
    session = {}

    # 统计表（必须）
    stats_file = request.files.get('stats_file')
    if not stats_file or not allowed_file(stats_file.filename):
        return jsonify({'ok': False, 'error': '请上传有效的统计表 (.xlsx)'}), 400

    stats_path = os.path.join(UPLOAD_DIR, f"{session_id}_stats.xlsx")
    stats_file.save(stats_path)
    session['stats_file'] = stats_path

    # 模板文件（可选，不传则使用默认模板）
    template_file = request.files.get('template_file')
    if template_file and allowed_file(template_file.filename):
        template_path = os.path.join(UPLOAD_DIR, f"{session_id}_template.xlsx")
        template_file.save(template_path)
        session['template_file'] = template_path
    else:
        default_template = os.path.join(PROJECT_ROOT, '模板.xlsx')
        if os.path.exists(default_template):
            session['template_file'] = default_template
        else:
            return jsonify({'ok': False, 'error': '未上传模板文件，且项目根目录下未找到 模板.xlsx'}), 400

    # 解析统计表
    students = parse_stats_file(stats_path)
    if not students:
        return jsonify({'ok': False, 'error': '统计表解析失败，请确认文件格式正确'}), 400

    # 格式化为前端可展示的列表
    student_list = []
    for idx, data in students.items():
        student_list.append({
            'index': idx,
            'name': data.get('data1', '未知姓名'),
            'education': data.get('data2', ''),
            'undergraduate': data.get('data3', ''),
            'first_choice_school': data.get('data4', ''),
            'first_choice_major': data.get('data5', ''),
            'score': data.get('data8', ''),
            'english': data.get('data9', ''),
            'politics': data.get('data10', ''),
            'transfer_type': data.get('data16', ''),
            'transfer_region': data.get('data18', ''),
        })

    session['students_raw'] = students
    session['student_list'] = student_list
    session['created_at'] = datetime.now().isoformat()
    sessions[session_id] = session

    return jsonify({
        'ok': True,
        'session_id': session_id,
        'count': len(student_list),
        'students': student_list,
    })


@app.route('/api/start', methods=['POST'])
def start_analysis():
    """启动 AI 分析任务（后台线程执行，通过 WebSocket 推送进度）"""
    body = request.get_json()
    session_id = body.get('session_id')
    selected_indices = body.get('indices', [])  # 选中的考生 index 列表

    session = sessions.get(session_id)
    if not session:
        return jsonify({'ok': False, 'error': '会话不存在或已过期'}), 400

    template_path = session['template_file']
    students_raw = session['students_raw']

    # 筛选选中的考生
    selected_students = {idx: students_raw[idx] for idx in selected_indices if idx in students_raw}
    if not selected_students:
        return jsonify({'ok': False, 'error': '未选择任何考生'}), 400

    # 为本次任务创建专属输出目录
    task_id = str(uuid.uuid4())[:8]
    task_output_dir = os.path.join(OUTPUT_DIR, task_id)
    os.makedirs(task_output_dir, exist_ok=True)

    # 启动后台线程
    def run_analysis():
        total = len(selected_students) * 3  # 每人 3 步
        step = 0

        results = []  # [{name, filename, success, error}]

        for idx, data in selected_students.items():
            name = data.get('data1', '未知姓名')

            # 步骤 1：准备数据
            step += 1
            socketio.emit('progress', {
                'task_id': task_id,
                'step': step, 'total': total,
                'name': name, 'phase': 'prepare',
                'message': f'正在准备 {name} 的数据...',
            })

            data_analysis, student_name = build_data_analysis(data)

            # 步骤 2：AI 分析
            step += 1
            socketio.emit('progress', {
                'task_id': task_id,
                'step': step, 'total': total,
                'name': name, 'phase': 'ai',
                'message': f'正在对 {name} 进行 AI 分析...',
            })

            try:
                ai_results = ai_tool.main(str(data_analysis))
                data_analysis += ai_results
                ai_success = True
            except Exception as e:
                ai_results = ['AI 分析失败，请稍后重试。', 'AI 分析失败，请稍后重试。']
                data_analysis += ai_results
                ai_success = False
                print(f"[ERROR] AI 分析失败 ({name}): {e}")

            # 步骤 3：生成 Excel
            step += 1
            socketio.emit('progress', {
                'task_id': task_id,
                'step': step, 'total': total,
                'name': name, 'phase': 'excel',
                'message': f'正在生成 {name} 的评估表...',
            })

            filename = f"{student_name}_xx调剂定制班1v1个人评估表.xlsx"
            output_path = os.path.join(task_output_dir, filename)
            success, error = generate_excel(data_analysis, template_path, output_path)

            results.append({
                'name': student_name,
                'filename': filename,
                'success': success and ai_success,
                'error': error,
            })

            socketio.emit('student_done', {
                'task_id': task_id,
                'name': student_name,
                'filename': filename,
                'success': success and ai_success,
                'error': error,
            })

        # 全部完成
        socketio.emit('task_done', {
            'task_id': task_id,
            'total': len(selected_students),
            'success': sum(1 for r in results if r['success']),
            'failed': sum(1 for r in results if not r['success']),
            'results': results,
        })

        # 保存任务信息到 session
        session['tasks'] = session.get('tasks', {})
        session['tasks'][task_id] = {
            'output_dir': task_output_dir,
            'results': results,
            'created_at': datetime.now().isoformat(),
        }

    threading.Thread(target=run_analysis, daemon=True).start()

    return jsonify({'ok': True, 'task_id': task_id})


@app.route('/api/download/<task_id>/<filename>')
def download_file(task_id, filename):
    """下载单个生成的 Excel 文件"""
    session = None
    for s in sessions.values():
        if 'tasks' in s and task_id in s['tasks']:
            session = s
            break

    if not session:
        return jsonify({'ok': False, 'error': '任务不存在'}), 404

    task_info = session['tasks'][task_id]
    filepath = os.path.join(task_info['output_dir'], filename)

    if not os.path.exists(filepath):
        return jsonify({'ok': False, 'error': '文件不存在'}), 404

    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/download_all/<task_id>')
def download_all(task_id):
    """将所有生成的文件打包为 ZIP 下载"""
    session = None
    for s in sessions.values():
        if 'tasks' in s and task_id in s['tasks']:
            session = s
            break

    if not session:
        return jsonify({'ok': False, 'error': '任务不存在'}), 404

    task_info = session['tasks'][task_id]
    output_dir = task_info['output_dir']

    zip_path = os.path.join(OUTPUT_DIR, f'{task_id}_all.zip')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in os.listdir(output_dir):
            if f.endswith('.xlsx'):
                zf.write(os.path.join(output_dir, f), f)

    return send_file(zip_path, as_attachment=True, download_name=f'评估表批量下载_{task_id}.zip')


# ═══════════════════════════════════════════════════════════════════════
#  启动
# ═══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print("=" * 50)
    print("  xx调剂定制班 - Web 服务")
    print("  访问 http://localhost:5000")
    print("=" * 50)
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, allow_unsafe_werkzeug=True)
